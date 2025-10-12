from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import pandas as pd
from os.path import join as path_join



class ExcelWriter:
    def __init__(self, competition_title, chief_secretary, chief_judge, protocols_dir):
        self.competition_title = competition_title
        self.participants_per_page = 26
        self.partners_per_page = self.participants_per_page // 2
        self.main_judge = chief_judge
        self.main_secretary = chief_secretary
        self.data_dir = protocols_dir

    def write_style(self, df, title="My Spanning Header"):
        wb = Workbook()
        ws = wb.active
        df.reset_index(drop=True, inplace=True)
        # Create a thin black border
        header_font = Font(bold=True, size=13)
        pages_total = df.shape[0] // self.participants_per_page + 1
        df_index = 0
        footer_idx0 = 3 * self.partners_per_page + 4
        footer_idx = footer_idx0
        for page in range(pages_total):
            index0 = (3 * self.partners_per_page + 6)*page
            index1 = index0 + 1
            index2 = index0 + 2
            ws[f'A{index1}'].font = header_font
            ws[f'A{index1}'].alignment = Alignment(horizontal='center', wrapText=True)
            ws[f'A{index1}'] = self.competition_title
            ws.row_dimensions[1].height = 65
            ws[f'A{index2}'].font = header_font
            ws[f'A{index2}'].alignment = Alignment(horizontal='center', wrapText=True)
            ws[f'A{index2}'] = title if pages_total == 1 else title + f"(стор. {page+1})"
            ws.row_dimensions[1].height = 65

            # Merge cells from A1 to C1 (inclusive)
            ws.merge_cells(f'A{index1}:D{index1}')
            ws.merge_cells(f'A{index2}:D{index2}')
            #
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))


            for half_row in range(index0//3, self.partners_per_page+index0//3):
                # Draw a diagonal line (top-left to bottom-right)
                ws[f'B{3 * half_row + 4}'].border = thin_border
                ws[f'C{3 * half_row + 4}'].border = thin_border
                # idx = 2*half_row
                if df_index in df.index:
                    ws[f'B{3 * half_row + 4}'].value = df.loc[df_index, "Команда"]
                    ws[f'C{3 * half_row + 4}'].value = df.loc[df_index, "Учасник"]
                ws[f'A{3 * half_row + 4}'] = f"{df_index+1} aka"
                df_index += 1
                ws[f'B{3 * half_row + 5}'].border = thin_border
                ws[f'C{3 * half_row + 5}'].border = thin_border
                # idx = 2*half_row+1
                if df_index in df.index:
                    ws[f'B{3 * half_row + 5}'].value = df.loc[df_index, "Команда"]
                    ws[f'C{3 * half_row + 5}'].value = df.loc[df_index, "Учасник"]
                ws[f'A{3 * half_row + 5}'] = f"{df_index+1} sira"
                df_index += 1
                ws[f'D{3*half_row+5}'].border = Border(diagonalUp=True, diagonal=Side(style='thin'))
                ws[f'D{3*half_row+4}'].border = Border(diagonalDown=True, diagonal=Side(style='thin'))
                ws[f'E{3 * half_row + 4}'].border = Border(bottom=Side(style='thin'))


            ws[f'B{footer_idx}'].border = Border(bottom=Side(style='thin'))
            ws[f'E{footer_idx}'].border = Border(bottom=Side(style='thin'))
            ws[f'B{footer_idx}'].value = self.main_judge
            ws[f'E{footer_idx}'].value = self.main_secretary
            ws[f'B{footer_idx + 1}'].value = "Головний Суддя"
            ws[f'E{footer_idx + 1}'].value = "Головний Секрктар"
            ws.row_dimensions[footer_idx + 1].height = 20

            footer_idx += footer_idx0+2
            # break
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 20
        wb.save(path_join(self.data_dir, f"{title}.xlsx"))

#
# # Create a new workbook and select the active sheet
# workbook = Workbook()
# sheet = workbook.active
#
# # Define your headers
# headers = ["Column A Header", "Column B Header", "Column C Header", "Column D Header"]
#
# # Define the starting row for your headers
# start_row = 1  # For example, to place headers in the first row
#
# # Write headers to the specified row
# for col_num, header_text in enumerate(headers, start=1):
#     sheet.cell(row=start_row, column=col_num, value=header_text)
#
# # If you need to write headers in another row as well (e.g., row 3)
# start_row_two = 3
# headers_two = ["Sub-Header 1", "Sub-Header 2", "Sub-Header 3", "Sub-Header 4"]
#
# for col_num, header_text in enumerate(headers_two, start=1):
#     sheet.cell(row=start_row_two, column=col_num, value=header_text)
#
# # Save the workbook
# # workbook.save("headers_example.xlsx")
#
# from openpyxl import Workbook
#
# # Create a new workbook and select the active sheet
# wb = Workbook()
# ws = wb.active
#
# # Set the value for the header in the top-left cell of the merged area
# ws['A1'] = "My Spanning Header"
#
# # Merge cells from A1 to C1 (inclusive)
# ws.merge_cells('A1:C1')
#
# # Optional: Apply styling to the merged header (e.g., bold, center alignment)
# from openpyxl.styles import Font, Alignment
# ws['A1'].font = Font(bold=True)
# ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
#
# # Add some data below the merged header for demonstration
# ws['A2'] = "Column 1 Data"
# ws['B2'] = "Column 2 Data"
# ws['C2'] = "Column 3 Data"
#
# # Save the workbook
# wb.save("merged_header_example.xlsx")

if __name__ == "__main__":
    df = pd.read_csv("/home/lityk/OlimpSecretary/data/test.csv")
    title = "Ката Хлопці (категорія A,  10-11 років)"
    obj = ExcelWriter(20, "Чемпіонат м Києва з карате")
    obj.write_style(df, title)
